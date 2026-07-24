#!/usr/bin/env python3
"""Incorporação de NOVOS cadastros: pessoas dos arquivos locais (CPF completo) que
NÃO existem no painel1_servidores. Anti-join por CPF (ix_p1_cpf), insert com
origem/nasc_fonte='arq:<caminho>'. Só CPF de 11 dígitos entra (sem CPF = sem dedup seguro)."""
import io, json, os, re, shutil, time, unicodedata

JOB = os.path.dirname(os.path.abspath(__file__))
TMP = os.path.join(JOB, "tmp"); os.makedirs(TMP, exist_ok=True)
LEDGER = os.path.join(JOB, "incorporados.jsonl")
RAIZ = "/Users/israelsantiago/Library/CloudStorage/GoogleDrive-israel.taptos@gmail.com/Outros computadores/Meu laptop/D:/"
import os
DB = os.environ["SUPABASE_DB_URL"]  # exporte de TCEMG/dados-painel/.supabase_env
CAP = 80 * 1024 * 1024
FLUSH_N = 100_000

def log(m): print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)

def norm(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.upper()).strip()

MUNIS = json.load(open(os.path.join(JOB, "municipios_br.json"), encoding="utf-8"))
IBGE_OK = {m["ibge"] for m in MUNIS}

def parse_date(v):
    from datetime import datetime, date
    if v is None: return None
    if isinstance(v, (datetime, date)):
        return f"{v.year:04d}-{v.month:02d}-{v.day:02d}" if 1900 <= v.year <= 2012 else None
    s = str(v).strip()
    for rx, order in ((r"^(\d{2})[\/\-\.](\d{2})[\/\-\.](\d{4})", "dmy"),
                      (r"^(\d{4})-(\d{2})-(\d{2})", "ymd"),
                      (r"^((?:19|20)\d{2})(\d{2})(\d{2})$", "ymd")):
        m = re.match(rx, s)
        if m:
            y, mo, d = (m.group(3), m.group(2), m.group(1)) if order == "dmy" else (m.group(1), m.group(2), m.group(3))
            if 1900 <= int(y) <= 2012 and 1 <= int(mo) <= 12 and 1 <= int(d) <= 31:
                return f"{y}-{mo}-{d}"
            return None
    return None

def path_ibge(path):
    for m7 in re.findall(r"\b(\d{7})\b", path):
        if m7 in IBGE_OK: return m7
    return None

def rows_iter(path, ext):
    if ext == ".csv":
        for enc in ("utf-8-sig", "latin-1"):
            try:
                with open(path, encoding=enc) as f:
                    first = f.readline()
                delim = max([";", ",", "\t", "|"], key=lambda d: first.count(d))
                import csv as _csv
                f = open(path, encoding=enc, newline="")
                yield from _csv.reader(f, delimiter=delim)
                f.close(); return
            except UnicodeDecodeError:
                continue
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            yield list(row)
        wb.close()

RX_NA = re.compile(r"NASC|ANIVERS"); RX_ID = re.compile(r"IDADE|FAIXA")
RX_CPF = re.compile(r"\bCPF\b|NR[_ ]?CPF|NUM[_ ]?CPF|NUM[_ ]?DOCUMENTO")
RX_NOME = re.compile(r"\bNOME\b|NM[_ ]?(SERVIDOR|CANDIDATO|PESSOA|FUNCIONARIO)|NOM[_ ]?PESSOA|NOME[_ ]?(COMPLETO|SERVIDOR|CIVIL)")
RX_SEXO = re.compile(r"\bSEXO\b|IND[_ ]?SEXO|GENERO")
RX_EMAIL = re.compile(r"E-?MAIL")
RX_TEL = re.compile(r"TELEFONE|\bFONE\b|CELULAR")
def val_sexo(v):
    v = norm(v)
    if v in ("M","MASCULINO","MASC"): return "M"
    if v in ("F","FEMININO","FEM"): return "F"
    return None

def pick(hn, rx, excl=None):
    for j, h in enumerate(hn):
        if rx.search(h) and (not excl or not excl.search(h)): return j
    return None

import psycopg2
def conecta():
    c = psycopg2.connect(DB, keepalives=1, keepalives_idle=30, keepalives_interval=10, keepalives_count=6)
    c.autocommit = False
    k = c.cursor(); k.execute("set statement_timeout=0"); c.commit()
    return c, k
conn, cur = conecta()

feitos = set()
if os.path.exists(LEDGER):
    for ln in open(LEDGER):
        try: feitos.add(json.loads(ln)["path"])
        except Exception: pass

novos = {}          # cpf -> dict(nome,nasc,ibge,fonte,sexo,email,tel)
pend_ledger = []    # entradas gravadas só após flush persistir
tot_ins = 0; sem_cpf = 0; tot_sx = tot_nx = tot_ex = tot_tx = 0

def flush():
    global conn, cur
    for tent in (1, 2):
        try:
            return _flush_db()
        except (psycopg2.OperationalError, psycopg2.DatabaseError) as e:
            log(f"conexão caiu no flush (tentativa {tent}): {str(e)[:80]} — reconectando")
            try: conn.close()
            except Exception: pass
            time.sleep(5)
            conn, cur = conecta()
    raise RuntimeError("flush falhou após reconexão")

def _flush_db():
    global novos, tot_ins
    if not novos: return 0
    cur.execute("""create temp table t_new (cpf text primary key, nome text, nasc date,
                   ibge text, fonte text, sexo text, email text, tel text) on commit drop""")
    buf = io.StringIO()
    def cl(x): return str(x).replace("\t"," ").replace("\\"," ").replace("\n"," ") if x else None
    for cpf, d in novos.items():
        buf.write("\t".join([cpf, cl(d['nome']) or '', d['nasc'] or '\\N', d['ibge'] or '\\N',
                             d['fonte'], d.get('sexo') or '\\N', cl(d.get('email')) or '\\N',
                             cl(d.get('tel')) or '\\N']) + "\n")
    buf.seek(0)
    cur.copy_expert("copy t_new from stdin with (format text, null '\\N')", buf)
    cur.execute("""
      insert into painel1_servidores (cpf, nome, ibge, data_nascimento, nasc_fonte, sexo, sexo_fonte,
                                      email, telefone, origem, consta_site)
      select t.cpf, t.nome, t.ibge, t.nasc,
             case when t.nasc is not null then t.fonte end,
             t.sexo, case when t.sexo is not null then t.fonte end,
             t.email, t.tel, t.fonte, 'nao'
        from t_new t
       where not exists (select 1 from painel1_servidores p where p.cpf = t.cpf)""")
    n = cur.rowcount
    cur.execute("""update painel1_servidores p set sexo=t.sexo, sexo_fonte=t.fonte
                    from t_new t where p.cpf=t.cpf and t.sexo is not null and p.sexo is null""")
    global tot_sx; tot_sx += cur.rowcount
    cur.execute("""update painel1_servidores p set data_nascimento=t.nasc, nasc_fonte=t.fonte
                    from t_new t where p.cpf=t.cpf and t.nasc is not null and p.data_nascimento is null""")
    global tot_nx; tot_nx += cur.rowcount
    cur.execute("""update painel1_servidores p set email=t.email
                    from t_new t where p.cpf=t.cpf and t.email is not null and (p.email is null or p.email='')""")
    global tot_ex; tot_ex += cur.rowcount
    cur.execute("""update painel1_servidores p set telefone=t.tel
                    from t_new t where p.cpf=t.cpf and t.tel is not null and (p.telefone is null or p.telefone='')""")
    global tot_tx; tot_tx += cur.rowcount
    conn.commit()
    tot_ins += n
    log(f"FLUSH: {len(novos)} candidatos → {n} NOVOS inseridos (acum {tot_ins})")
    novos = {}
    global pend_ledger
    if pend_ledger:
        with open(LEDGER, "a") as lf:
            for e in pend_ledger: lf.write(json.dumps(e, ensure_ascii=False) + "\n")
        pend_ledger = []
    return n

doc = json.load(open(os.path.join(JOB, "candidatos.json")))
fila = [c for c in doc["candidatos"] if c["path"].replace(RAIZ, "") not in feitos and c.get("birth_cols")]
log(f"{len(fila)} arquivos na fila de incorporação")
for c in fila:
    path, ext, sz = c["path"], c["ext"], c["size"]
    rel = path.replace(RAIZ, "")
    r = {"path": rel, "status": "ok", "cpfs": 0}
    if sz > CAP:
        r["status"] = "grande_pulado"
    else:
        loc = os.path.join(TMP, "inc" + ext)
        try:
            shutil.copyfile(path, loc)
            it = rows_iter(loc, ext)
            hdr = next(it, None)
            if hdr is None:
                r["status"] = "vazio"
            else:
                hn = [norm(h) for h in hdr]
                jc, jm = pick(hn, RX_CPF), pick(hn, RX_NOME)
                jn = pick(hn, RX_NA, RX_ID)
                js, je, jt = pick(hn, RX_SEXO), pick(hn, RX_EMAIL), pick(hn, RX_TEL)
                ibp = path_ibge(path)
                if jc is None or jm is None:
                    r["status"] = "sem_cpf_ou_nome"
                else:
                    fonte = ("arq:" + rel)[:120]
                    for row in it:
                        if jc >= len(row) or jm >= len(row): continue
                        cpf = re.sub(r"\D", "", str(row[jc] or ""))
                        if len(cpf) != 11 or cpf == "00000000000":
                            sem_cpf += 1; continue
                        nome = norm(row[jm])
                        if len(nome) < 5: continue
                        nasc = parse_date(row[jn]) if (jn is not None and jn < len(row)) else None
                        sx = val_sexo(row[js]) if (js is not None and js < len(row)) else None
                        em = (str(row[je]).strip() or None) if (je is not None and je < len(row) and row[je] and '@' in str(row[je])) else None
                        tl = (str(row[jt]).strip() or None) if (jt is not None and jt < len(row) and row[jt]) else None
                        d = novos.get(cpf)
                        if d is None:
                            novos[cpf] = {'nome': nome, 'nasc': nasc, 'ibge': ibp, 'fonte': fonte,
                                          'sexo': sx, 'email': em, 'tel': tl}
                            r["cpfs"] += 1
                        else:
                            if nasc and not d['nasc']: d['nasc'] = nasc
                            if sx and not d.get('sexo'): d['sexo'] = sx
                            if em and not d.get('email'): d['email'] = em
                            if tl and not d.get('tel'): d['tel'] = tl
            os.remove(loc)
        except Exception as e:
            try: conn.rollback()
            except Exception: pass
            r = {"path": rel, "status": "erro", "erro": str(e)[:150]}
    pend_ledger.append(r)
    if len(novos) >= FLUSH_N:
        flush()
    time.sleep(0.5)
flush()
log(f"passe 1 (arquivos): {tot_ins} novos · updates existentes: sexo {tot_sx} · nasc {tot_nx} · email {tot_ex} · tel {tot_tx} · sem CPF: {sem_cpf}")

# ---- passe 2 · Painel 6: responsáveis fora do P1 (sem CPF → dedup nome+município) ----
cur.execute("""
  insert into painel1_servidores (nome, cpf, ibge, orgao, email, responsabilidade, origem, consta_site)
  select r.nome, case when r.cpf ~ '^[0-9]{11}$' then r.cpf end,
         r.cod_ibge::text, r.orgao, r.email, r.tipo_responsabilidade,
         'p6:'||coalesce(nullif(r.origem,''),'resp'), 'nao'
    from (select distinct on (f_unaccent(upper(nome)), cod_ibge) *
            from painel6_responsaveis where coalesce(no_painel1,'') = 'Não' and nome is not null) r
   where not exists (select 1 from painel1_servidores p
                     where f_unaccent(upper(p.nome)) = f_unaccent(upper(r.nome))
                       and p.ibge = r.cod_ibge::text)""")
n_p6 = cur.rowcount; conn.commit()
log(f"passe 2 (Painel 6 → novos): {n_p6}")

# ---- passe 3 · CRM: pessoas criadas pela equipe, fora do P1 ----
cur.execute("""
  insert into painel1_servidores (nome, cpf, ibge, orgao, setor, cargo_funcao, email, telefone, origem, consta_site)
  select s.nome, case when s.cpf ~ '^[0-9]{11}$' then s.cpf end,
         e.cod_ibge::text, e.nome, st.nome, s.cargo, s.email, s.telefone,
         'crm:'||coalesce(nullif(s.origem,''),'manual'), 'nao'
    from crm_servidores s
    left join crm_entidades e on e.id = s.entidade_id
    left join crm_setores st on st.id = s.setor_id
   where s.painel1_id is null and coalesce(s.origem,'') <> 'painel1' and s.nome is not null
     and not exists (select 1 from painel1_servidores p
                     where (s.cpf ~ '^[0-9]{11}$' and p.cpf = s.cpf)
                        or (f_unaccent(upper(p.nome)) = f_unaccent(upper(s.nome))
                            and p.ibge = e.cod_ibge::text))""")
n_crm = cur.rowcount; conn.commit()
log(f"passe 3 (CRM → novos): {n_crm}")

# ---- passe 4 · emails do Painel 6 p/ registros existentes (por CPF, só onde vazio) ----
cur.execute("""
  update painel1_servidores p
     set email = r.email
    from (select distinct on (cpf) cpf, email from painel6_responsaveis
           where cpf ~ '^[0-9]{11}$' and email is not null and email <> '') r
   where p.cpf = r.cpf and (p.email is null or p.email = '')""")
n_em = cur.rowcount; conn.commit()
log(f"passe 4 (emails P6 por CPF): {n_em}")

cur.execute("select count(*) from painel1_servidores")
log(f"CONSOLIDAÇÃO CONCLUÍDA · arquivos:{tot_ins} + P6:{n_p6} + CRM:{n_crm} novos · emails:{n_em} · base agora: {cur.fetchone()[0]}")
conn.close()
