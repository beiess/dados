#!/usr/bin/env python3
"""Scanner v2: cabeçalhos de CSV/XLSX (FUSE, lotes) com CHECKPOINT e FLUSH incremental
de candidatos.json — o motor de enriquecimento consome em paralelo."""
import json, os, re, time, unicodedata

JOB = os.path.dirname(os.path.abspath(__file__))
XLSX_MAX = 30 * 1024 * 1024
LOTE, PAUSA = 50, 0.5

def norm(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.upper()).strip()

RX_NASC = re.compile(r"NASC|ANIVERS")
RX_IDADE = re.compile(r"\bIDADE\b|FAIXA[_ ]?ETARIA")
RX = {"cpf": re.compile(r"\bCPF\b|NR[_ ]?CPF|NUM[_ ]?CPF"),
      "matricula": re.compile(r"MATRICULA|\bMASP\b"),
      "nome": re.compile(r"\bNOME\b|NM[_ ]?(SERVIDOR|CANDIDATO|PESSOA|FUNCIONARIO)|NOME[_ ]?(COMPLETO|SERVIDOR|CIVIL)"),
      "municipio": re.compile(r"MUNICIPIO|CIDADE|NM[_ ]?UE|\bIBGE\b|COD[_ ]?IBGE"),
      "uf": re.compile(r"^UF$|SG[_ ]?UF|ESTADO"),
      "orgao": re.compile(r"ORGAO|ENTIDADE|UNIDADE[_ ]?GESTORA|LOTACAO")}

def csv_header(path):
    for enc in ("utf-8-sig", "latin-1"):
        try:
            with open(path, "r", encoding=enc) as f:
                chunk = f.read(16384)
            line = chunk.splitlines()[0] if chunk else ""
            if not line:
                return None
            delim = max([";", ",", "\t", "|"], key=lambda d: line.count(d))
            if line.count(delim) == 0:
                return [line.strip()]
            return [c.strip().strip('"') for c in line.split(delim)]
        except UnicodeDecodeError:
            continue
        except Exception:
            return None
    return None

def xlsx_header(path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            wb.close()
            return [str(c) if c is not None else "" for c in row]
        wb.close()
    except Exception:
        return None
    return None

lista = []
for ln in open(os.path.join(JOB, "lista_prio.txt"), encoding="utf-8", errors="replace"):
    ln = ln.rstrip("\n")
    if "|" in ln:
        sz, p = ln.split("|", 1)
        lista.append((int(sz), p))

CK = os.path.join(JOB, "scan_ck.txt")
OUT = os.path.join(JOB, "candidatos.json")
start = int(open(CK).read().strip()) if os.path.exists(CK) else 0
doc = json.load(open(OUT)) if os.path.exists(OUT) and start > 0 else {"candidatos": [], "pulados": [], "meta": {}}

def flush(i, done=False):
    doc["meta"] = {"total": len(lista), "pos": i, "done": done, "ts": time.strftime("%H:%M:%S")}
    tmp = OUT + ".tmp"
    json.dump(doc, open(tmp, "w", encoding="utf-8"), ensure_ascii=False)
    os.replace(tmp, OUT)
    open(CK, "w").write(str(i))

print(f"{len(lista)} arquivos · retomando de {start}", flush=True)
t0 = time.time()
for i in range(start, len(lista)):
    sz, path = lista[i]
    if i % LOTE == 0 and i > start:
        flush(i)
        time.sleep(PAUSA)
        print(f"  [{i}/{len(lista)}] {time.time()-t0:.0f}s · candidatos={len(doc['candidatos'])}", flush=True)
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls") and sz > XLSX_MAX:
        doc["pulados"].append({"path": path, "size": sz, "motivo": "xlsx grande"})
        continue
    hdr = csv_header(path) if ext == ".csv" else xlsx_header(path)
    if not hdr:
        continue
    hn = [norm(h) for h in hdr]
    birth = [hdr[j] for j, h in enumerate(hn) if RX_NASC.search(h)]
    idade = [hdr[j] for j, h in enumerate(hn) if RX_IDADE.search(h)]
    if not birth and not idade:
        continue
    keys = {k: [hdr[j] for j, h in enumerate(hn) if rx.search(h)] for k, rx in RX.items()}
    doc["candidatos"].append({"path": path, "size": sz, "ext": ext,
                              "birth_cols": birth, "idade_cols": idade,
                              "keys": {k: v for k, v in keys.items() if v}, "cols": hdr[:40]})
flush(len(lista), done=True)
print(f"SCAN CONCLUÍDO: {len(doc['candidatos'])} candidatos · {len(doc['pulados'])} pulados · {time.time()-t0:.0f}s", flush=True)
