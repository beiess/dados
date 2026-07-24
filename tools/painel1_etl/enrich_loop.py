#!/usr/bin/env python3
"""Motor de enriquecimento (Fase 3): consome candidatos.json conforme o scanner avança.
Por arquivo: copia p/ disco local (≤80MB), extrai (chave→nascimento), atualiza
painel1_servidores.data_nascimento SÓ ONDE NULO, nasc_fonte='arq:<caminho relativo>'.
Chaves: CPF (prioridade) e nome+município (data única; ibge por coluna, nome de
município+UF, ou inferido do CAMINHO). Ledger em processed.jsonl."""
import io, json, os, re, shutil, sys, time, unicodedata
from datetime import datetime, date

JOB = os.path.dirname(os.path.abspath(__file__))
TMP = os.path.join(JOB, "tmp"); os.makedirs(TMP, exist_ok=True)
LEDGER = os.path.join(JOB, "processed.jsonl")
RAIZ = "/Users/israelsantiago/Library/CloudStorage/GoogleDrive-israel.taptos@gmail.com/Outros computadores/Meu laptop/D:/"
import os
DB = os.environ["SUPABASE_DB_URL"]  # exporte de TCEMG/dados-painel/.supabase_env
CAP = 80 * 1024 * 1024
MIN_PARES = 10

def log(m): print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)

def norm(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.upper()).strip()

MUNIS = json.load(open(os.path.join(JOB, "municipios_br.json"), encoding="utf-8"))
BY_UF = {}
for m in MUNIS:
    BY_UF.setdefault(m["uf"], {})[norm(m["nome"])] = m["ibge"]
IBGE_OK = {m["ibge"] for m in MUNIS}
UFS = set(BY_UF)

def parse_date(v):
    if v is None: return None
    if isinstance(v, (datetime, date)):
        y = v.year
        return f"{y:04d}-{v.month:02d}-{v.day:02d}" if 1900 <= y <= 2012 else None
    s = str(v).strip()
    m = re.match(r"^(\d{2})[\/\-\.](\d{2})[\/\-\.](\d{4})", s)
    if m: d, mo, y = m.group(1), m.group(2), m.group(3)
    else:
        m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
        if m: y, mo, d = m.group(1), m.group(2), m.group(3)
        else:
            m = re.match(r"^((?:19|20)\d{2})(\d{2})(\d{2})$", s)  # AAAAMMDD (SICOM)
            if not m: return None
            y, mo, d = m.group(1), m.group(2), m.group(3)
    if not (1900 <= int(y) <= 2012 and 1 <= int(mo) <= 12 and 1 <= int(d) <= 31): return None
    return f"{y}-{mo}-{d}"

def path_muni(path):
    """Infere (ibge) do caminho: código IBGE de 7 dígitos ou UF+nome de município nos segmentos."""
    for m7 in re.findall(r"\b(\d{7})\b", path):
        if m7 in IBGE_OK:
            return m7
    segs = [norm(s) for s in path.replace("\\", "/").split("/")]
    uf = next((s for s in segs if s in UFS), None)
    cand = []
    for s in segs:
        s2 = re.sub(r"[_\-]", " ", s)
        s2 = re.sub(r"\.(CSV|XLSX|XLS)$", "", s2)
        if uf and s2 in BY_UF[uf]: cand.append(BY_UF[uf][s2])
        elif not uf:
            hits = [BY_UF[u][s2] for u in ("MG",) if s2 in BY_UF[u]]
            cand += hits
    return cand[0] if len(set(cand)) == 1 else None

def rows_iter(path, ext):
    if ext == ".csv":
        for enc in ("utf-8-sig", "latin-1"):
            try:
                with open(path, encoding=enc) as f:
                    first = f.readline()
                delim = max([";", ",", "\t", "|"], key=lambda d: first.count(d))
                import csv as _csv
                f = open(path, encoding=enc, newline="")
                rd = _csv.reader(f, delimiter=delim)
                yield from rd
                f.close()
                return
            except UnicodeDecodeError:
                continue
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            yield list(row)
        wb.close()

def pick(hn, rx, excl=None):
    for j, h in enumerate(hn):
        if rx.search(h) and (not excl or not excl.search(h)):
            return j
    return None

RX_NA = re.compile(r"NASC|ANIVERS"); RX_ID = re.compile(r"IDADE|FAIXA")
RX_CPF = re.compile(r"\bCPF\b|NR[_ ]?CPF|NUM[_ ]?CPF|NUM[_ ]?DOCUMENTO")
RX_NOME = re.compile(r"\bNOME\b|NM[_ ]?(SERVIDOR|CANDIDATO|PESSOA|FUNCIONARIO)|NOME[_ ]?(COMPLETO|SERVIDOR|CIVIL)")
RX_MUN = re.compile(r"MUNICIPIO|CIDADE|NM[_ ]?UE"); RX_IBGE = re.compile(r"IBGE")
RX_UF = re.compile(r"^UF$|SG[_ ]?UF")

def processa(c, cur):
    path, ext, sz = c["path"], c["ext"], c["size"]
    rel = path.replace(RAIZ, "")
    if not c.get("birth_cols"):
        return {"path": rel, "status": "so_idade"}
    if sz > CAP:
        return {"path": rel, "status": "grande_pulado", "size": sz}
    loc = os.path.join(TMP, "cur" + ext)
    try:
        shutil.copyfile(path, loc)
    except Exception as e:
        return {"path": rel, "status": "erro_copia", "erro": str(e)[:120]}
    it = rows_iter(loc, ext)
    try:
        hdr = next(it)
    except StopIteration:
        return {"path": rel, "status": "vazio"}
    hn = [norm(h) for h in hdr]
    jn = pick(hn, RX_NA, RX_ID)
    if jn is None:
        return {"path": rel, "status": "sem_col_nasc"}
    jc, jm = pick(hn, RX_CPF), pick(hn, RX_NOME)
    jmu, jib, juf = pick(hn, RX_MUN), pick(hn, RX_IBGE), pick(hn, RX_UF)
    ib_path = path_muni(path)
    cpf_map, nome_map, lidas = {}, {}, 0
    for row in it:
        lidas += 1
        if lidas > 2_000_000: break
        if jn >= len(row): continue
        dt = parse_date(row[jn])
        if not dt: continue
        if jc is not None and jc < len(row):
            cpf = re.sub(r"\D", "", str(row[jc] or ""))
            if len(cpf) == 11 and cpf != "00000000000":
                if cpf in cpf_map and cpf_map[cpf] != dt: cpf_map[cpf] = "X"
                else: cpf_map.setdefault(cpf, dt)
        if jm is not None and jm < len(row):
            nm = norm(row[jm])
            if len(nm) >= 8:
                ib = None
                if jib is not None and jib < len(row):
                    d7 = re.sub(r"\D", "", str(row[jib] or ""))[:7]
                    ib = d7 if d7 in IBGE_OK else None
                if not ib and jmu is not None and jmu < len(row):
                    uf = norm(row[juf]) if (juf is not None and juf < len(row)) else None
                    tab = BY_UF.get(uf) if uf in UFS else None
                    if tab: ib = tab.get(norm(row[jmu]))
                    else:
                        hits = [BY_UF[u].get(norm(row[jmu])) for u in ("MG",)]
                        ib = hits[0] if hits and hits[0] else None
                if not ib: ib = ib_path
                if ib:
                    k = (nm, ib)
                    if k in nome_map and nome_map[k] != dt: nome_map[k] = "X"
                    else: nome_map.setdefault(k, dt)
    cpf_map = {k: v for k, v in cpf_map.items() if v != "X"}
    nome_map = {k: v for k, v in nome_map.items() if v != "X"}
    os.remove(loc)
    if len(cpf_map) + len(nome_map) < MIN_PARES:
        return {"path": rel, "status": "poucos_pares", "linhas": lidas,
                "cpf": len(cpf_map), "nome": len(nome_map)}
    fonte = ("arq:" + rel)[:120]
    up_c = up_n = 0
    if cpf_map:
        cur.execute("create temp table t_c (cpf text primary key, nasc date) on commit drop")
        buf = io.StringIO(); [buf.write(f"{k}\t{v}\n") for k, v in cpf_map.items()]; buf.seek(0)
        cur.copy_expert("copy t_c from stdin", buf)
        cur.execute("""update painel1_servidores p set data_nascimento=c.nasc, nasc_fonte=%s
                       from t_c c where p.cpf=c.cpf and p.data_nascimento is null""", (fonte,))
        up_c = cur.rowcount
    if nome_map:
        cur.execute("create temp table t_n (nome_norm text, ibge text, nasc date, primary key(nome_norm,ibge)) on commit drop")
        buf = io.StringIO(); [buf.write(f"{k[0]}\t{k[1]}\t{v}\n") for k, v in nome_map.items()]; buf.seek(0)
        cur.copy_expert("copy t_n from stdin", buf)
        cur.execute("""update painel1_servidores p set data_nascimento=c.nasc, nasc_fonte=%s
                       from t_n c where p.ibge=c.ibge and f_unaccent(upper(p.nome))=c.nome_norm
                         and p.data_nascimento is null""", (fonte,))
        up_n = cur.rowcount
    cur.connection.commit()
    return {"path": rel, "status": "ok", "linhas": lidas, "pares_cpf": len(cpf_map),
            "pares_nome": len(nome_map), "upd_cpf": up_c, "upd_nome": up_n}

import psycopg2
conn = psycopg2.connect(DB); conn.autocommit = False
cur = conn.cursor(); cur.execute("set statement_timeout=0"); conn.commit()

feitos = set()
if os.path.exists(LEDGER):
    for ln in open(LEDGER):
        try: feitos.add(json.loads(ln)["path"])
        except Exception: pass

tot_c = tot_n = 0
while True:
    try:
        doc = json.load(open(os.path.join(JOB, "candidatos.json")))
    except Exception:
        time.sleep(30); continue
    fila = [c for c in doc["candidatos"] if c["path"].replace(RAIZ, "") not in feitos]
    fila.sort(key=lambda c: (0 if c.get("keys", {}).get("cpf") else 1, -c["size"]))
    if not fila:
        if doc["meta"].get("done"):
            break
        time.sleep(45); continue
    for c in fila:
        rel = c["path"].replace(RAIZ, "")
        try:
            r = processa(c, cur)
        except Exception as e:
            conn.rollback()
            r = {"path": rel, "status": "erro", "erro": str(e)[:200]}
        feitos.add(rel)
        open(LEDGER, "a").write(json.dumps(r, ensure_ascii=False) + "\n")
        if r.get("upd_cpf") or r.get("upd_nome"):
            tot_c += r.get("upd_cpf", 0); tot_n += r.get("upd_nome", 0)
            log(f"UPD +{r.get('upd_cpf',0)}/cpf +{r.get('upd_nome',0)}/nome ← {rel[-70:]}")
        elif r["status"] not in ("ok",):
            log(f"{r['status']} ← {rel[-70:]}")
        time.sleep(1)

cur.execute("select count(*) from painel1_servidores where data_nascimento is not null")
log(f"ENRIQUECIMENTO CONCLUÍDO · novos: {tot_c} por CPF + {tot_n} por nome · total na base: {cur.fetchone()[0]}")
conn.close()
