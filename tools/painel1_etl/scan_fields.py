#!/usr/bin/env python3
"""Inventário de CAMPOS complementares: lê cabeçalhos de TODOS os CSV/XLSX (FUSE, lotes,
checkpoint) e persiste headers.jsonl + fields_inventory.json — que categorias de dado
(email/telefone/cargo/sexo/endereço/redes/matrícula/remuneração/vínculo/escolaridade)
existem em cada arquivo e com que chave (CPF/nome/matrícula/município)."""
import json, os, re, time, unicodedata

JOB = os.path.dirname(os.path.abspath(__file__))
XLSX_MAX = 30 * 1024 * 1024
LOTE, PAUSA = 50, 0.5
HDRS = os.path.join(JOB, "headers.jsonl")
CK = os.path.join(JOB, "fields_ck.txt")

def norm(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.upper()).strip()

CATS = {
  "email": r"E-?MAIL",
  "telefone": r"TELEFONE|\bFONE\b|CELULAR|WHATS",
  "cargo": r"CARGO|FUNCAO|\bFUNC\b",
  "sexo": r"\bSEXO\b|GENERO|IND[_ ]?SEXO",
  "endereco": r"ENDERECO|LOGRADOURO|\bCEP\b|BAIRRO",
  "redes": r"INSTAGRAM|FACEBOOK|LINKEDIN|REDE[_ ]?SOCIAL|SITE|TWITTER",
  "matricula": r"MATRICULA|\bMASP\b",
  "remuneracao": r"REMUNERACAO|SALARIO|VENCIMENTO|PROVENTO|SUBSIDIO",
  "vinculo": r"ADMISSAO|EXERCICIO|SITUACAO|VINCULO|REGIME|LOTACAO",
  "escolaridade": r"ESCOLARIDADE|FORMACAO|GRAU[_ ]?INSTRUCAO",
  "nascimento": r"NASC|ANIVERS",
  "obito": r"OBITO|FALECIMENTO",
}
KEYS = {"cpf": r"\bCPF\b|NR[_ ]?CPF|NUM[_ ]?CPF|NUM[_ ]?DOCUMENTO",
        "nome": r"\bNOME\b|NM[_ ]?(SERVIDOR|CANDIDATO|PESSOA|FUNCIONARIO)|NOM[_ ]?PESSOA",
        "municipio": r"MUNICIPIO|CIDADE|\bIBGE\b",
        "orgao": r"ORGAO|ENTIDADE|LOTACAO|UNIDADE"}
CATS = {k: re.compile(v) for k, v in CATS.items()}
KEYS = {k: re.compile(v) for k, v in KEYS.items()}

def csv_header(path):
    for enc in ("utf-8-sig", "latin-1"):
        try:
            with open(path, "r", encoding=enc) as f:
                chunk = f.read(16384)
            line = chunk.splitlines()[0] if chunk else ""
            if not line:
                return None
            delim = max([";", ",", "\t", "|"], key=lambda d: line.count(d))
            if line.count(delim) == 0:
                return [line.strip()]
            return [c.strip().strip('"') for c in line.split(delim)]
        except UnicodeDecodeError:
            continue
        except Exception:
            return None
    return None

def xlsx_header(path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            wb.close()
            return [str(c) if c is not None else "" for c in row]
        wb.close()
    except Exception:
        return None
    return None

lista = []
for ln in open(os.path.join(JOB, "lista_prio.txt"), encoding="utf-8", errors="replace"):
    ln = ln.rstrip("\n")
    if "|" in ln:
        sz, p = ln.split("|", 1)
        lista.append((int(sz), p))

start = int(open(CK).read().strip()) if os.path.exists(CK) else 0
print(f"{len(lista)} arquivos · retomando de {start}", flush=True)
t0 = time.time()
out = open(HDRS, "a", encoding="utf-8")
for i in range(start, len(lista)):
    sz, path = lista[i]
    if i % LOTE == 0 and i > start:
        out.flush(); open(CK, "w").write(str(i))
        time.sleep(PAUSA)
        print(f"  [{i}/{len(lista)}] {time.time()-t0:.0f}s", flush=True)
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls") and sz > XLSX_MAX:
        out.write(json.dumps({"path": path, "size": sz, "skip": "grande"}) + "\n")
        continue
    hdr = csv_header(path) if ext == ".csv" else xlsx_header(path)
    if not hdr:
        out.write(json.dumps({"path": path, "size": sz, "skip": "ilegivel"}) + "\n")
        continue
    hn = [norm(h) for h in hdr]
    cats = {c: [hdr[j] for j, h in enumerate(hn) if rx.search(h)] for c, rx in CATS.items()}
    keys = {k: bool([1 for h in hn if rx.search(h)]) for k, rx in KEYS.items()}
    out.write(json.dumps({"path": path, "size": sz, "cols": hdr[:60],
                          "cats": {k: v for k, v in cats.items() if v},
                          "keys": {k: v for k, v in keys.items() if v}},
                         ensure_ascii=False) + "\n")
out.flush(); out.close(); open(CK, "w").write(str(len(lista)))

# consolida inventário
inv = {"por_categoria": {}, "arquivos_ricos": []}
regs = [json.loads(l) for l in open(HDRS, encoding="utf-8") if l.strip()]
vistos = {}
for r in regs:
    vistos[r["path"]] = r   # última leitura vence (dedup de reruns)
for r in vistos.values():
    if "cats" not in r: continue
    nc = r.get("cats", {}); ky = r.get("keys", {})
    for c in nc:
        e = inv["por_categoria"].setdefault(c, {"arquivos": 0, "com_cpf": 0, "com_nome": 0})
        e["arquivos"] += 1
        if ky.get("cpf"): e["com_cpf"] += 1
        if ky.get("nome"): e["com_nome"] += 1
    score = len(nc) + (2 if ky.get("cpf") else 0)
    if score >= 5:
        inv["arquivos_ricos"].append({"path": r["path"].split("D:/")[-1], "cats": list(nc), "keys": list(ky), "size": r["size"]})
inv["arquivos_ricos"].sort(key=lambda x: -len(x["cats"]))
inv["arquivos_ricos"] = inv["arquivos_ricos"][:80]
json.dump(inv, open(os.path.join(JOB, "fields_inventory.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("INVENTÁRIO DE CAMPOS CONCLUÍDO:", flush=True)
for c, e in sorted(inv["por_categoria"].items(), key=lambda kv: -kv[1]["arquivos"]):
    print(f"  {c}: {e['arquivos']} arquivos ({e['com_cpf']} c/ CPF, {e['com_nome']} c/ nome)", flush=True)
